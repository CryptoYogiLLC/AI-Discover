"""Export service for generating CSV, Excel, and JSON exports"""

import csv
import json
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from structlog import get_logger

from app.models.assessment import ApplicationAssessment
from app.models.project import Project

logger = get_logger()


class ExportService:
    """Service for exporting assessment data in various formats"""

    def __init__(self):
        self.field_metadata = self._get_field_metadata()

    async def export_assessments_csv(
        self, assessments: List[ApplicationAssessment], include_metadata: bool = True
    ) -> bytes:
        """
        Export assessments to CSV format

        Args:
            assessments: List of assessments to export
            include_metadata: Whether to include metadata columns

        Returns:
            CSV file as bytes
        """
        output = io.StringIO()

        # Define columns
        columns = self._get_csv_columns(include_metadata)

        writer = csv.DictWriter(output, fieldnames=columns, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()

        # Write assessment data
        for assessment in assessments:
            row = self._assessment_to_csv_row(assessment, include_metadata)
            writer.writerow(row)

        # Convert to bytes
        output.seek(0)
        return output.getvalue().encode("utf-8")

    async def export_assessments_excel(
        self,
        assessments: List[ApplicationAssessment],
        project: Optional[Project] = None,
        include_analytics: bool = True,
    ) -> bytes:
        """
        Export assessments to Excel format with multiple sheets

        Args:
            assessments: List of assessments to export
            project: Optional project information
            include_analytics: Whether to include analytics sheet

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, assessments, project)
        self._create_assessments_sheet(wb, assessments)
        self._create_details_sheet(wb, assessments)

        if include_analytics:
            self._create_analytics_sheet(wb, assessments)

        # Add metadata sheet
        self._create_metadata_sheet(wb, assessments)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    async def export_assessments_json(
        self,
        assessments: List[ApplicationAssessment],
        include_metadata: bool = True,
        pretty_print: bool = True,
    ) -> bytes:
        """
        Export assessments to JSON format

        Args:
            assessments: List of assessments to export
            include_metadata: Whether to include export metadata
            pretty_print: Whether to format JSON for readability

        Returns:
            JSON file as bytes
        """
        export_data = {
            "export_date": datetime.utcnow().isoformat(),
            "total_assessments": len(assessments),
            "assessments": [],
        }

        if include_metadata:
            export_data["metadata"] = {
                "version": "1.0",
                "fields": self.field_metadata,
                "export_type": "full",
            }

        # Convert assessments to dictionaries
        for assessment in assessments:
            assessment_dict = self._assessment_to_dict(assessment)
            export_data["assessments"].append(assessment_dict)

        # Convert to JSON
        if pretty_print:
            json_str = json.dumps(export_data, indent=2, default=str)
        else:
            json_str = json.dumps(export_data, default=str)

        return json_str.encode("utf-8")

    def _get_csv_columns(self, include_metadata: bool) -> List[str]:
        """Get CSV column names"""
        columns = [
            # Basic Information
            "id",
            "project_id",
            "application_name",
            "application_description",
            "business_owner",
            "technical_owner",
            # Business Criticality
            "business_criticality",
            "revenue_impact",
            "user_count",
            # Technical Architecture
            "architecture_type",
            "technology_stack",
            "programming_languages",
            # Infrastructure
            "server_count",
            "database_types",
            "storage_requirements_gb",
            # Integration
            "integration_count",
            "integration_types",
            "external_dependencies",
            # Performance
            "peak_load_users",
            "response_time_sla_ms",
            "availability_sla_percent",
            # Compliance
            "compliance_requirements",
            "data_sensitivity",
            "authentication_methods",
            # Current State
            "technical_debt_score",
            "code_quality_score",
            "documentation_quality",
            # Cloud Readiness
            "containerization_ready",
            "stateless_architecture",
            "cloud_native_services_used",
            # Migration Complexity
            "data_volume_gb",
            "transaction_volume_per_day",
            "batch_processing_required",
            "real_time_processing_required",
            # Costs
            "current_licensing_cost_annual",
            "infrastructure_cost_annual",
            "support_cost_annual",
            # Timeline
            "migration_deadline",
            # Risk
            "migration_risk_score",
            "business_impact_if_failed",
            "rollback_complexity",
            # Team
            "team_cloud_experience_score",
            "training_required",
            "external_support_needed",
            # Lifecycle
            "last_major_update",
            "expected_retirement_date",
            "modernization_planned",
            # Network
            "bandwidth_requirements_mbps",
            "latency_requirements_ms",
            "vpn_requirements",
            # Operational
            "support_hours",
            # Migration Approach
            "preferred_migration_strategy",
            "acceptable_downtime_hours",
            "data_migration_approach",
            # Cost-Benefit
            "expected_cost_savings_percent",
            "expected_performance_improvement_percent",
            "expected_roi_months",
            # Recommendations
            "recommendation",
            "recommendation_score",
            "recommendation_reasoning",
        ]

        if include_metadata:
            columns.extend(
                [
                    "assessment_date",
                    "assessed_by",
                    "last_updated",
                    "notes",
                ]
            )

        return columns

    def _assessment_to_csv_row(
        self, assessment: ApplicationAssessment, include_metadata: bool
    ) -> Dict[str, Any]:
        """Convert assessment to CSV row"""
        row = {}

        # Add all fields
        for column in self._get_csv_columns(include_metadata):
            value = getattr(assessment, column, None)

            # Handle special types
            if isinstance(value, (list, dict)):
                row[column] = json.dumps(value) if value else ""
            elif isinstance(value, datetime):
                row[column] = value.isoformat()
            elif isinstance(value, bool):
                row[column] = "Yes" if value else "No"
            elif value is None:
                row[column] = ""
            else:
                row[column] = str(value)

        return row

    def _assessment_to_dict(self, assessment: ApplicationAssessment) -> Dict[str, Any]:
        """Convert assessment to dictionary"""
        data = {}

        # Get all attributes
        for key in dir(assessment):
            if not key.startswith("_") and not callable(getattr(assessment, key)):
                value = getattr(assessment, key)

                # Handle special types
                if hasattr(value, "isoformat"):
                    data[key] = value.isoformat()
                elif hasattr(value, "__dict__"):
                    continue  # Skip relationships
                else:
                    data[key] = value

        return data

    def _create_summary_sheet(
        self,
        wb: Workbook,
        assessments: List[ApplicationAssessment],
        project: Optional[Project],
    ):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Application Assessment Export Summary"
        ws["A1"].font = Font(size=16, bold=True)

        # Project info
        row = 3
        if project:
            ws[f"A{row}"] = "Project:"
            ws[f"B{row}"] = project.name
            row += 1
            ws[f"A{row}"] = "Description:"
            ws[f"B{row}"] = project.description or "N/A"
            row += 2

        # Summary stats
        ws[f"A{row}"] = "Total Assessments:"
        ws[f"B{row}"] = len(assessments)
        row += 1

        # Recommendation breakdown
        recommendation_counts = {}
        for assessment in assessments:
            if assessment.recommendation:
                rec = assessment.recommendation.value
                recommendation_counts[rec] = recommendation_counts.get(rec, 0) + 1

        row += 1
        ws[f"A{row}"] = "Recommendations:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        for rec, count in recommendation_counts.items():
            ws[f"A{row}"] = rec
            ws[f"B{row}"] = count
            row += 1

        # Criticality breakdown
        criticality_counts = {}
        for assessment in assessments:
            if assessment.business_criticality:
                crit = assessment.business_criticality
                criticality_counts[crit] = criticality_counts.get(crit, 0) + 1

        row += 1
        ws[f"A{row}"] = "Business Criticality:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        for crit, count in criticality_counts.items():
            ws[f"A{row}"] = crit
            ws[f"B{row}"] = count
            row += 1

        # Export metadata
        row += 2
        ws[f"A{row}"] = "Export Date:"
        ws[f"B{row}"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _create_assessments_sheet(
        self, wb: Workbook, assessments: List[ApplicationAssessment]
    ):
        """Create main assessments sheet"""
        ws = wb.create_sheet("Assessments")

        # Headers
        headers = [
            "Application Name",
            "Business Owner",
            "Technical Owner",
            "Criticality",
            "Architecture",
            "Users",
            "Revenue Impact",
            "Recommendation",
            "Risk Score",
            "Technical Debt",
            "Cloud Ready",
            "Expected Savings %",
        ]

        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row, assessment in enumerate(assessments, 2):
            ws.cell(row=row, column=1, value=assessment.application_name)
            ws.cell(row=row, column=2, value=assessment.business_owner or "")
            ws.cell(row=row, column=3, value=assessment.technical_owner or "")
            ws.cell(row=row, column=4, value=assessment.business_criticality or "")
            ws.cell(row=row, column=5, value=assessment.architecture_type or "")
            ws.cell(row=row, column=6, value=assessment.user_count or 0)
            ws.cell(row=row, column=7, value=assessment.revenue_impact or 0)
            ws.cell(
                row=row,
                column=8,
                value=(
                    assessment.recommendation.value if assessment.recommendation else ""
                ),
            )
            ws.cell(row=row, column=9, value=assessment.migration_risk_score or 0)
            ws.cell(row=row, column=10, value=assessment.technical_debt_score or 0)
            ws.cell(
                row=row,
                column=11,
                value="Yes" if assessment.containerization_ready else "No",
            )
            ws.cell(
                row=row, column=12, value=assessment.expected_cost_savings_percent or 0
            )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def _create_details_sheet(
        self, wb: Workbook, assessments: List[ApplicationAssessment]
    ):
        """Create detailed assessments sheet"""
        ws = wb.create_sheet("Detailed Data")

        # Get all columns
        columns = self._get_csv_columns(include_metadata=True)

        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)

        # Write data
        for row_idx, assessment in enumerate(assessments, 2):
            row_data = self._assessment_to_csv_row(assessment, include_metadata=True)
            for col_idx, column in enumerate(columns, 1):
                value = row_data.get(column, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _create_analytics_sheet(
        self, wb: Workbook, assessments: List[ApplicationAssessment]
    ):
        """Create analytics sheet"""
        ws = wb.create_sheet("Analytics")

        # Title
        ws["A1"] = "Assessment Analytics"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3

        # Average scores
        ws[f"A{row}"] = "Average Metrics:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        # Calculate averages
        risk_scores = [
            a.migration_risk_score for a in assessments if a.migration_risk_score
        ]
        debt_scores = [
            a.technical_debt_score for a in assessments if a.technical_debt_score
        ]
        quality_scores = [
            a.code_quality_score for a in assessments if a.code_quality_score
        ]

        ws[f"A{row}"] = "Average Risk Score:"
        ws[f"B{row}"] = (
            round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else "N/A"
        )
        row += 1

        ws[f"A{row}"] = "Average Technical Debt:"
        ws[f"B{row}"] = (
            round(sum(debt_scores) / len(debt_scores), 2) if debt_scores else "N/A"
        )
        row += 1

        ws[f"A{row}"] = "Average Code Quality:"
        ws[f"B{row}"] = (
            round(sum(quality_scores) / len(quality_scores), 2)
            if quality_scores
            else "N/A"
        )
        row += 2

        # Cloud readiness
        ws[f"A{row}"] = "Cloud Readiness:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        container_ready = sum(1 for a in assessments if a.containerization_ready)
        stateless = sum(1 for a in assessments if a.stateless_architecture)

        ws[f"A{row}"] = "Container Ready:"
        ws[f"B{row}"] = (
            f"{container_ready} ({container_ready/len(assessments)*100:.1f}%)"
        )
        row += 1

        ws[f"A{row}"] = "Stateless Architecture:"
        ws[f"B{row}"] = f"{stateless} ({stateless/len(assessments)*100:.1f}%)"
        row += 2

        # Cost analysis
        ws[f"A{row}"] = "Cost Analysis:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        total_current_cost = sum(
            (a.current_licensing_cost_annual or 0)
            + (a.infrastructure_cost_annual or 0)
            + (a.support_cost_annual or 0)
            for a in assessments
        )

        total_savings = sum(
            (
                (a.current_licensing_cost_annual or 0)
                * (a.expected_cost_savings_percent or 0)
                / 100
            )
            for a in assessments
        )

        ws[f"A{row}"] = "Total Annual Cost:"
        ws[f"B{row}"] = f"${total_current_cost:,.2f}"
        row += 1

        ws[f"A{row}"] = "Projected Annual Savings:"
        ws[f"B{row}"] = f"${total_savings:,.2f}"

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_metadata_sheet(
        self, wb: Workbook, assessments: List[ApplicationAssessment]
    ):
        """Create metadata sheet"""
        ws = wb.create_sheet("Metadata")

        # Title
        ws["A1"] = "Export Metadata"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3

        # Export info
        ws[f"A{row}"] = "Export Information:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = "Export Date:"
        ws[f"B{row}"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        row += 1

        ws[f"A{row}"] = "Total Records:"
        ws[f"B{row}"] = len(assessments)
        row += 1

        ws[f"A{row}"] = "Export Format:"
        ws[f"B{row}"] = "Excel Workbook"
        row += 1

        ws[f"A{row}"] = "Version:"
        ws[f"B{row}"] = "1.0"
        row += 2

        # Field descriptions
        ws[f"A{row}"] = "Field Descriptions:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        for field, metadata in self.field_metadata.items():
            ws[f"A{row}"] = field
            ws[f"B{row}"] = metadata.get("description", "")
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 60

    def _get_field_metadata(self) -> Dict[str, Dict[str, Any]]:
        """Get metadata for all assessment fields"""
        return {
            "application_name": {
                "description": "Name of the application being assessed",
                "type": "string",
                "required": True,
            },
            "business_criticality": {
                "description": "Business criticality level (Critical, High, Medium, Low)",
                "type": "enum",
                "values": ["Critical", "High", "Medium", "Low"],
            },
            "revenue_impact": {
                "description": "Annual revenue impact in dollars",
                "type": "number",
                "unit": "USD",
            },
            "technical_debt_score": {
                "description": "Technical debt score on a scale of 1-10",
                "type": "integer",
                "min": 1,
                "max": 10,
            },
            "migration_risk_score": {
                "description": "Migration risk score on a scale of 1-10",
                "type": "integer",
                "min": 1,
                "max": 10,
            },
            "recommendation": {
                "description": "6R migration recommendation",
                "type": "enum",
                "values": [
                    "rehost",
                    "replatform",
                    "repurchase",
                    "refactor",
                    "retire",
                    "retain",
                ],
            },
            # Add more field metadata as needed
        }
